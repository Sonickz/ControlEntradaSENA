from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import authenticate, login, logout
from django.core.paginator import Paginator
from django.db.models import Prefetch, Q
from django.http import HttpResponse
from django.contrib import messages
from .models import *
from .forms import *
from openpyxl import Workbook #Generar archivos excel
from io import BytesIO
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


#Funcion para crear paginacion segun un modelo
def createPagination(request, name, model, cant):
    paginator = Paginator(model, cant)  # Paginator
    page = request.GET.get(name) # Pagina actual
    pages = paginator.get_page(page) # Enviar el paginator y detectar la actual
    return pages

#=============================================================

#Login admin
@user_passes_test(lambda u: not u.is_authenticated, login_url='adminpanel') #Si el usuario ya esta logeado no va a poder acceder a la pagina de login 
def login_admin(request):  # sourcery skip: use-named-expression

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        admin = authenticate(request, username=username, password=password)
        
        if admin:
            login(request, admin)
            return redirect('adminpanel')
        else:
            messages.error(request, "Usuario o contraseña incorrectos")

    return render(request, 'login.html', {
        'title': 'Iniciar Sesion',
    })

#Logout admin
def logout_admin(request):
    logout(request)
    return redirect('login')

#=============================================================

#AdminPanel
@login_required(login_url="login")
def adminpanel(request):
    return render(request, 'adminpanel.html')

#Ingresos
@login_required(login_url="admin")
def access(request):
    access = Ingresos.objects.all().exclude(idingreso__in=Salidas.objects.values('ingreso')) #Ingresos activos
    access = createPagination(request, "access", access, 100) #Paginacion

    exits = Salidas.objects.all() #Salidas
    exits = createPagination(request, "exits", exits, 100) #Paginacion
    
    searchAccess = None
    searchExit = None
    search = request.GET.get("search")
    if search:
        searchAccess = Ingresos.objects.filter(
            Q(idingreso__icontains=search) | Q(fecha__icontains=search) | Q(usuario__nombres__icontains=search) |
            Q(usuario__apellidos__icontains=search) | Q(usuario__documento__icontains=search) |
            Q(vehiculo__placa__icontains=search) | Q(horaingreso__icontains=search)
        )
        searchAccess= createPagination(request, "access-search", searchAccess, 100)
        
        searchExit = Salidas.objects.filter(
            Q(idsalida__icontains=search) |  Q(fecha__icontains=search) | 
            Q(ingreso__usuario__nombres__icontains=search) | Q(ingreso__usuario__apellidos__icontains=search) | 
            Q(ingreso__usuario__documento__icontains=search) |
            Q(vehiculo__placa__icontains=search) | Q(ingreso__horaingreso__icontains=search) |
            Q(horasalida__icontains=search)
        )
        searchExit = createPagination(request, "exit-search", searchExit, 100)
        
                
    return render(request, "pages/ingresos/access.html", {
        'title': 'Ingresos',
        'access': access,
        'exits': exits,
        'search_access': searchAccess,
        'search_exit': searchExit
    })

#Lista usuarios
@login_required(login_url="admin")
def users(request):
    
    users = {}
    roles = Roles.objects.all()
    for rol in roles:
        model = Usuarios.objects.filter(rol=rol.idrol)
        rol_nombre = rol.nombre.lower()
        users[rol_nombre] = {
                "name": rol_nombre,
                "model": createPagination(request, f"{rol_nombre}", model, 100)
        }

    search = request.GET.get("search")
    if search:
        search = Usuarios.objects.filter(
            Q(nombres__icontains=search) | Q(apellidos__icontains=search) | Q(documento__icontains=search) |
            Q(telefono__icontains=search) | Q(correo__icontains=search) | Q(centro__nombre__icontains=search) |
            Q(ficha__numero__icontains=search) | Q(rol__nombre__icontains=search)
        )
        search = createPagination(request, "user-search", search, 100)
               
    return render(request, 'pages/usuarios/users.html', {
        'title': 'Usuarios',
        'roles': roles,
        'users': users,
        'search': search
        })

#Registrar usuario
def register_user(request, rol):
    initial = {'rol': rol}
    rol_selected = Roles.objects.get(idrol=rol)
    form = RegisterUser(request.POST or None, request.FILES or None, initial=initial)

    form.fields['centro'].required = rol != 3
    form.fields['ficha'].required = rol == 2
    form.fields['documento'].widget.attrs['readonly'] = False

    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Se ha registrado correctamente")
        return redirect('users')

    return render(request, 'pages/usuarios/register.html', {
        'title': 'Registrar usuario',
        'form': form,
        'rol': rol,
        'rol_selected': rol_selected
    })

#Editar usuario
def edit_user(request, id):
    instance = Usuarios.objects.get(idusuario=id)
    rol = str(instance.rol.idrol)
    rol_selected = Roles.objects.get(idrol=rol)
    form = RegisterUser(request.POST or None, request.FILES or None, instance=instance)

    form.fields['centro'].required = instance.rol != 3
    form.fields['ficha'].required = instance.rol == 2
    form.fields['rol'].required = False

    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Se ha editado correctamente")
        return redirect('users')

    return render(request, 'pages/usuarios/edit.html', {
        'title': 'Editar usuario',
        'form': form,
        'rol': rol,
        'rol_selected': rol_selected
    })

#=============================================================

#Vehiculos
@login_required(login_url="admin")
def vehicles(request):
    vehiculos = Vehiculos.objects.all()
    vehiculos = createPagination(request, "vehicle", vehiculos, 100)

    search = request.GET.get("search")
    if search:
        search = Vehiculos.objects.filter(
            Q(idvehiculo__icontains=search) | Q(usuario__nombres__icontains=search) | Q(usuario__apellidos__icontains=search) | Q(usuario__documento__icontains=search) |
            Q(tipo__nombre__icontains=search) | Q(placa__icontains=search) | Q(marca__nombre__icontains=search) | Q(modelo__icontains=search)
        )
        search = createPagination(request, "vehicle-search", search, 100)

    return render(request, 'pages/vehiculos/vehicles.html', {
        'title': 'Vehiculos',
        'vehiculos': vehiculos,
        'search': search
    })

#Editar vehiculos
def edit_vehicle(request, id):
    instance=Vehiculos.objects.get(idvehiculo=id)
    vehicle = str(instance.tipo.idtipovehiculo)
    form = RegisterVehicle(request.POST or None, request.FILES or None, instance= instance)
    
    if request .method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Vehículo editado correctamente")
        return redirect('vehiculos')
            
    return render(request, 'pages/vehiculos/edit.html', {
         'title': 'Editar Vehiculos',
         'vehicle': vehicle,
         'form': form 
    })

#=============================================================

# Lista de dispositivos
@login_required(login_url="admin")
def devices(request):
    
    dispositivos = Dispositivos.objects.all()
    dispositivos = createPagination(request, "device", dispositivos, 100)

    search = request.GET.get("search")
    if search:
        search = Dispositivos.objects.filter(
            Q(iddispositivo__icontains=search) | Q(usuario__nombres__icontains=search) | Q(usuario__apellidos__icontains=search) | Q(usuario__documento__icontains=search) | 
            Q(sn__icontains=search) | Q(tipo__nombre__icontains=search) | Q(marca__nombre__icontains=search)
        )
        search = createPagination(request, "device-search", search, 100)
    
    return render(request, 'pages/dispositivos/devices.html', {
        'title': 'Dispositivos',
        'dispositivos': dispositivos,
        'search': search
    })

#Editar dispositivo
def edit_device(request, id):
    instance = Dispositivos.objects.get(iddispositivo=id)
    form = RegisterDevice(request.POST or None, request.FILES or None, instance= instance)

    if request .method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "se ha editado correctamente")
        return redirect('dispositivo')
        
    return render(request, 'pages/dispositivos/edit.html', {
        'title': 'Editar dispositivo',
        'form': form
    })

#=============================================================================

#acerda de
@login_required(login_url="admin")
def about(request):
    
    return render(request, 'pages/acerca/about.html', {
        'title': 'Acerca de'
    })

#Registrar admin
@login_required(login_url="admin")
def register_admin(request):
    form = RegisterAdmin(request.POST or None)

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.info(request, "success-admin")
        return redirect('adminpanel')

    return render(request, 'pages/registeradmin.html', {
        'title': "Registrar administrador",
        'form': form
    })
#===========================================================================
#Reporte en excel
def reports(request, model):

    #Crear libro
    wb = Workbook()    
    #Crear hoja
    ws_sheet = wb.active
        
    

    #Reporte de Ingresos y salidas
    if model == "ingresos":
        name = "Ingresos y Salidas"
        reportData(ws_sheet, model) 
        ws_sheet2 = wb.create_sheet(title="Salidas")
        reportData(ws_sheet2, model="salidas")

    elif model == "usuarios":
        name = "Usuarios" 
        reportData(ws_sheet, model="aprendiz")
        ws_sheet2 = wb.create_sheet(title="Instructores")
        reportData(ws_sheet2, model="instructor")
        ws_sheet3 = wb.create_sheet(title="Visitantes")
        reportData(ws_sheet3, model="visitante")
        ws_sheet4 = wb.create_sheet(title="Administrativos")
        reportData(ws_sheet4, model="administrativo")

    else:
        name = reportData(ws_sheet, model)

    
    
        

    # Crear un objeto BytesIO para guardar el archivo en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Configurar la respuesta HTTP para descargar el archivo
    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename=Reporte de {name} - ControlEntradaSENA V1.1.0.xlsx"

    return response

def reportData(ws_sheet, model):
    if model == "ingresos":
        model = Ingresos.objects.all().prefetch_related("ingresosdispositivos_set")
        ws_sheet.title = "Ingresos" #Titulo hoja
        #Columnas
        ws_sheet.append(["Id", "Fecha", "Usuario", "Vehiculo", "Dispositivos", "Hora Ingreso"]) 
        header_cells = ws_sheet["A1:F1"]
        #Llenar hoja
        for ingreso in model: 
            dispositivos = []     
            for ingreso_dispositivo in ingreso.ingresosdispositivos_set.all():                
                dispositivos.append(str(ingreso_dispositivo.dispositivo))
            ws_sheet.append([ingreso.idingreso,
                    ingreso.fecha,
                    f"{str(ingreso.usuario)}: {ingreso.usuario.documento}",
                    str(ingreso.vehiculo) if ingreso.vehiculo else "Ninguno",
                    ", ".join(dispositivos) if dispositivos else "Ninguno",
                    ingreso.horaingreso])                                                  
            
    elif model == "salidas":
        model = Salidas.objects.all().prefetch_related("salidasdispositivos_set")
        ws_sheet.title = "Salidas" 
        ws_sheet.append(["Id", "Fecha", "Ingreso", "Usuario", "Vehiculo", "Dispositivos", "Hora Ingreso", "Hora Salida" ]) 
        header_cells = ws_sheet["A1:H1"]
        for salida in model:
            dispositivos = []
            for salida_dispositivo in salida.salidasdispositivos_set.all():
                dispositivos.append(str(salida_dispositivo.dispositivo))
            ws_sheet.append([salida.idsalida,
                                 salida.fecha,
                                 salida.ingreso.idingreso,
                                f"{str(salida.ingreso.usuario)}: {salida.ingreso.usuario.documento}",
                                str(salida.vehiculo)if salida.vehiculo else "Ninguno",
                                ", ".join(dispositivos) if dispositivos else "Ninugno",
                                salida.ingreso.horaingreso,
                                salida.horasalida])
            
    elif model == "aprendiz":
        model = Usuarios.objects.filter(rol=2).prefetch_related("vehiculos_set").prefetch_related("dispositivos_set")        
        ws_sheet.title = "Aprendices" 
        ws_sheet.append(["Id", "Nombres", "Apellidos", "Tipo de Documento", "Documento", "Telefono", "Correo", "Centro", "Rol", "Ficha", "Dispositivos", "Vehiculos"]) 
        header_cells = ws_sheet["A1:L1"]
        for user in model:
            dispositivos = []
            vehiculos = []
            for dispositivo in user.dispositivos_set.all():
                dispositivos.append(str(dispositivo))
            for vehiculo in user.vehiculos_set.all():
                vehiculos.append(str(vehiculo))
            ws_sheet.append([user.idusuario,
                                user.nombres,
                                user.apellidos,
                                str(user.tipodocumento),
                                user.documento,
                                user.telefono,
                                user.correo,
                                str(user.centro),
                                str(user.rol),
                                str(user.ficha),
                                ", ".join(dispositivos) if dispositivos else "Ninguno",
                                ", ".join(vehiculos) if vehiculos else "Ninguno"
                                ])
            
    elif model == "instructor" or model == "administrativo":
        if model == "instructor":
            model = Usuarios.objects.filter(rol=1).prefetch_related("vehiculos_set").prefetch_related("dispositivos_set")
            ws_sheet.title = "Instructores" 
        else:
            model = Usuarios.objects.filter(rol=4).prefetch_related("vehiculos_set").prefetch_related("dispositivos_set")        
            ws_sheet.title = "Administradores" 
            
        ws_sheet.append(["Id", "Nombres", "Apellidos", "Tipo de Documento", "Documento", "Telefono", "Correo", "Centro", "Rol", "Dispositivos", "Vehiculos"]) 
        header_cells = ws_sheet["A1:K1"]
        for user in model:
            dispositivos = []
            vehiculos = []
            for dispositivo in user.dispositivos_set.all():
                dispositivos.append(str(dispositivo))
            for vehiculo in user.vehiculos_set.all():
                vehiculos.append(str(vehiculo))
            ws_sheet.append([user.idusuario,
                                user.nombres,
                                user.apellidos,
                                str(user.tipodocumento),
                                user.documento,
                                user.telefono,
                                user.correo,
                                str(user.centro),
                                str(user.rol),
                                ", ".join(dispositivos) if dispositivos else "Ninguno",
                                ", ".join(vehiculos) if vehiculos else "Ninguno"
                                ])
            
    elif model == "visitante":
        model = Usuarios.objects.filter(rol=3).prefetch_related("vehiculos_set").prefetch_related("dispositivos_set")        
        ws_sheet.title = "Visitantes" 
        ws_sheet.append(["Id", "Nombres", "Apellidos", "Tipo de Documento", "Documento", "Telefono", "Correo", "Rol", "Dispositivos", "Vehiculos"]) 
        header_cells = ws_sheet["A1:J1"]
        for user in model:
            dispositivos = []
            vehiculos = []
            for dispositivo in user.dispositivos_set.all():
                dispositivos.append(str(dispositivo))
            for vehiculo in user.vehiculos_set.all():
                vehiculos.append(str(vehiculo))
            ws_sheet.append([user.idusuario,
                                user.nombres,
                                user.apellidos,
                                str(user.tipodocumento),
                                user.documento,
                                user.telefono,
                                user.correo,
                                str(user.rol),
                                ", ".join(dispositivos) if dispositivos else "Ninguno",
                                ", ".join(vehiculos) if vehiculos else "Ninguno"
                                ])
            
    elif model == "dispositivos":
        model = Dispositivos.objects.all()
        ws_sheet.title = "Dispositivos"
        ws_sheet.append(["Id", "Usuario", "SN", "Tipo", "Marca"])
        header_cells = ws_sheet["A1:E1"]
        for dispositivo in model:
            ws_sheet.append([dispositivo.iddispositivo,
                                f"{str(dispositivo.usuario)}: {dispositivo.usuario.documento}",
                                dispositivo.sn,
                                str(dispositivo.tipo),
                                str(dispositivo.marca)])
            
    elif model == "vehiculos":
        model = Vehiculos.objects.all()
        ws_sheet.title = "Vehiculos"
        ws_sheet.append(["Id", "Usuario", "Placa", "Tipo", "Marca", "Modelo"])
        header_cells = ws_sheet["A1:F1"]
        for vehiculo in model:
            ws_sheet.append([vehiculo.idvehiculo,
                                f"{str(vehiculo.usuario)}: {vehiculo.usuario.documento}",
                                vehiculo.placa,
                                str(vehiculo.tipo),
                                str(vehiculo.marca),
                                vehiculo.modelo])
    
    designExcel(header_cells, ws_sheet)
    
    return ws_sheet.title         
        

def designExcel(header_cells, ws_sheet):
    #Diseño de celdas
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), 
                         right=Side(style="thin"), 
                         top=Side(style="thin"), 
                         bottom=Side(style="thin"))

    #Establecer diseño de encabezado
    for row in header_cells:
        for cell in row:
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.fill = header_fill
            cell.border = thin_border            
    
    # Establecer alineacion de celdas
    for row in ws_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = center_alignment
            if cell.value:                
                cell.border = thin_border

    # Establecer ancho de columnas
    column_width = 40
    for column in ws_sheet.columns:
        ws_sheet.column_dimensions[column[0].column_letter].width = column_width
        

#==========================================================================

from django.shortcuts import render

def informacion_view(request):
    # Lógica para la vista de información de la aplicación
    return render(request, 'pages/acerca/informacion.html')

def integrantes_view(request):
    # Lógica para la vista de integrantes del grupo
    return render(request, 'pages/acerca/integrantes.html')

def lideres_view(request):
    # Lógica para la vista de líderes del proyecto
    return render(request, 'pages/acerca/lideres.html')


